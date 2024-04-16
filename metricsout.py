import sys
import matplotlib.pyplot as plt
from openpyxl.workbook import Workbook
import openpyxl
import openpyxl.styles.borders as borders
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Border


def write_results(arr):
    sys.stdout = open(f"out/reddit_{arr[4]}cores.txt", "w")
    print("Vertex Count:", arr[0])
    print("Edge count: %d" % arr[1])
    print(f"Algorithm execution time: {arr[2]} seconds")
    print(f"Algorithm and data parsing execution time: {arr[3]} seconds")
    print("Core Count %d" % arr[4])
    print("Number of extend operations: %d" % arr[5])
    print("Number of non-extended operations: %d" % arr[6])
    print("Min Sum Square: %d, Initial: %d" % (arr[7], arr[8]))
    print("Min Sum Square without delta: %d" % arr[9])
    print("Highest Volume %d, Initial: %d" % (arr[10], arr[11]))
    sys.stdout.close()


def save_basic_plot(x, y1, xlabel, ylabel, title):
    fig, ax = plt.subplots()
    ax.plot(x, y1)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    plt.savefig(f"out/{title}.png")


def save_2basic_plot(x, y1, y2, xlabel, ylabel, title, y1name, y2name, yscale=None):
    fig, ax = plt.subplots()
    ax.plot(x, y1, label=y1name)
    ax.plot(x, y2, label=y2name)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    if yscale is not None:
        ax.set_yscale(yscale)
    ax.legend()
    plt.savefig(f"out/{title}.png")


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):
        col_name = openpyxl.utils.cell.get_column_letter(min_col + i)
        value = column_widths[i] + 1
        ws.column_dimensions[col_name].width = value


def get_algorithm_name(args):
    name = ""
    if not args.noconstructive:
        name = "const"
    if not args.noiterative:
        if len(name) > 0:
            name += " and "
        name += "iter"
    return name


def create_excel(args, data, datasets):
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("other")
    DEFAULT_FONT.sz = 10
    # _font = Font(sz=10)
    # {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    ws["A1"] = "Dataset"
    interval = len(data) / len(datasets)

    ws["B1"] = "Vertex\nCnt"
    ws["C1"] = "Edge\nCnt"
    ws["D1"] = "Time (s)"
    ws["D2"] = "Algo"
    ws["E2"] = "Total"
    ws["F1"] = "Cores"
    ws["G1"] = "Reassignment"
    ws["G2"] = "Cnt"
    ws["H2"] = "Non-R. Cnt"
    ws["I2"] = "Vol"
    # now ws worksheet data:
    ws["J1"] = "Improvement(%)"
    ws["J2"] = "Delay"
    ws["K2"] = "Max"
    ws["L1"] = "Vol"
    ws["L2"] = "Avg"
    ws["M2"] = "I CV"
    ws["N2"] = "O CV"
    ws["O2"] = "I Max"
    ws["P2"] = "O Max"
    ws["Q2"] = "I Min"
    ws["R2"] = "O Min"
    ws["S1"] = "Degree"
    ws["S2"] = "Cnt"
    ws["T2"] = "Avg"
    ws["U2"] = "Total"
    ws["V2"] = "Max"
    ws["W2"] = "Min"
    ws["X2"] = "CV"

    for i in range(len(data)):
        d = data[i]
        increment = 2
        row_idx = i + 3
        for j in range(len(d)):
            if j == 9:
                ws.cell(row=row_idx, column=j + increment, value=f"=ROUND(100*(H{row_idx}-I{row_idx})/I{row_idx}, 3)")
                increment += 1
            ws.cell(row=row_idx, column=j + increment, value=d[j])
    double = borders.Side(border_style="thin")

    # Get the range of cells to copy
    cells_to_copy = ws["A:I"]

    # Iterate over each row in the range
    for row in cells_to_copy:
        # Iterate over each cell in the row
        for cell in row:
            # Create a new cell in the second worksheet with the same value as the original cell
            ws1[cell.coordinate] = cell.value
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=17)
    ws1.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws1.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws1.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    ws1.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
    ws1.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    # Now remove B-I from ws except F
    ws.delete_cols(2, 4)
    ws.delete_cols(3, 3)
    for i in range(ws.max_column):
        ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(1)].alignment = openpyxl.styles.Alignment(
            horizontal="center")
        for j in range(1, ws.max_row):
            cell = ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(j + 1)]
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="right", vertical="center")
            if j % interval == 1:
                cell.border = Border(bottom=double)
        adjust_column_width_from_col(ws, 2, 1, ws.max_column)
        adjust_column_width_from_col(ws, 1, 1, 1)

    # for j in range(1, ws.max_row + 1):
    #     end_cell = ws.cell(row=j, column=ws.max_column)
    #     end_cell.border = Border(right=double, bottom=double)
    # row = ws.row_dimensions[2]
    for worksheet in [ws, ws1]:
        for i in range(len(datasets)):
            d = datasets[i]
            worksheet.merge_cells(start_row=interval * i + 3, start_column=1, end_row=interval * (i + 1) + 2,
                                  end_column=1)
            cell = worksheet.cell(row=interval * i + 3, column=1, value=d)
            cell.border = Border(right=double)

    # Save the workbook
    wb.save(f'out/{args.core_cnt}cores-{args.alpha}-{get_algorithm_name(args)}-node{args.node_core_count}.xlsx')


def print_vol_excel(vol_df: dict[str, list[int]], node_core_cnt, core_cnt) -> None:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"] = "Dataset"
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws["B1"] = "Out of Node Vol Send"
    ws["B2"] = "Max"
    ws["C2"] = "Min"
    ws["D2"] = "Avg"
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    ws["E1"] = "Total Vol Send"
    ws["E2"] = "Max"
    ws["F2"] = "Min"
    ws["G2"] = "Avg"
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=10)
    ws["H1"] = "Out of Node Vol Recv"
    ws["H2"] = "Max"
    ws["I2"] = "Min"
    ws["J2"] = "Avg"
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
    ws["K1"] = "Total Vol Recv"
    ws["K2"] = "Max"
    ws["L2"] = "Min"
    ws["M2"] = "Avg"
    i = 3
    for name, vlist in vol_df.items():
        ws.cell(row=i, column=1, value=name)
        for j, v in enumerate(vlist):
            ws.cell(row=i, column=2 + j, value=v)
        i += 1

    for i in range(ws.max_column):
        ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(1)].alignment = openpyxl.styles.Alignment(
            horizontal="center")
        for j in range(1, ws.max_row):
            cell = ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(j + 1)]
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="right", vertical="center")
    adjust_column_width_from_col(ws, 2, 1, ws.max_column)
    adjust_column_width_from_col(ws, 1, 1, 1)
    wb.save(f'out/vol-{len(vol_df)}-{core_cnt}-{node_core_cnt}.xlsx')
