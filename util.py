import math
from enum import Enum

import numpy as np
from openpyxl.styles import Border

from torch_geometric.datasets import Planetoid, PPI, Reddit, Amazon, KarateClub, AmazonProducts, Yelp, Flickr
import torch_geometric.transforms as transforms
import torch_geometric.utils as tl
import pymetis
from scipy.sparse import coo_matrix
import sys
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.workbook import Workbook
import openpyxl
import openpyxl.styles.borders as borders
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import Font


def cv(x):
    return np.std(x, ddof=1) / np.mean(x) * 100


def get_coo_mat(dataset_name):
    match dataset_name:
        case "Reddit":
            dataset = Reddit(root="data/Reddit")
        case "Amazon":
            dataset = Amazon(root="data/Amazon", name="Computers")
        case "AmazonProducts":
            dataset = AmazonProducts(root="data/AmazonProducts")
        case "KarateClub":
            dataset = KarateClub()
        case "Yelp":
            dataset = Yelp(root="data/Yelp")
        case "Flickr":
            dataset = Flickr(root="data/Flickr")
        case _:
            raise "invalid dataset"
    data = dataset[0]
    coo_data = data.edge_index.numpy()
    vtx_count = data.num_nodes
    # print("Directed:", data.is_directed())
    mappings = [[] for _ in range(vtx_count)]
    for i in range(len(coo_data[0])):
        mappings[coo_data[0][i]].append(coo_data[1][i])
    return coo_data, vtx_count, mappings


def create_coo_mat():
    mat = np.zeros((2, 14), dtype=int)
    mat[0, :] = [1, 1, 2, 2, 3, 4, 4, 5, 6, 6, 7, 7, 7, 7]
    mat[1, :] = [2, 7, 4, 6, 2, 1, 6, 5, 4, 5, 1, 3, 4, 7]
    mat -= 1
    return mat, 7


def get_volume(d: dict[int, list[int]]):
    return sum([len(x) for x in d.values()])


def calc_square(arr: list[dict]):
    ret = 0
    for d in arr:
        ret += np.sum(len(v) for v in d.values()) ** 2
    return ret


# used for testing purposes
def count_maps(maps: (int, dict)):
    return np.sum(len(v) for v in maps[1].values())


def get_uniform_mapping(cpu_cnt, vtx_count):
    return np.array([i % cpu_cnt for i in range(vtx_count)])


class DestData:
    def __init__(self, vid):
        self.expands = dict()
        self.reassign_cores = dict()
        self.volume = 0
        self.id = vid

    # calculate the cost from the dict itself rather than the volume variable
    def cost_raw(self):
        vol = sum(len(v) for v in self.expands.values())
        return vol ** 2

    def cost(self):
        return self.volume ** 2

    def insert(self, key, values):
        if key not in self.expands:
            self.expands[key] = set()
        first_len = len(self.expands[key])

        self.expands[key].update(values)
        self.expands[key].discard(self.id)
        self.volume += len(self.expands[key]) - first_len

    def set_forwarded_core(self, key, core_id):
        # if key in self.reassign_cores:
        #     raise Exception("Double forwarding attempt!")
        self.reassign_cores[key] = core_id

    def sqr_contribution_of(self, key):
        return self.cost() - (self.volume - len(self.expands[key])) ** 2

    def remove_assignment(self, key):
        e_len = len(self.expands[key])
        self.reassign_cores.pop(key, None)
        self.expands.pop(key)
        self.volume -= e_len
        return e_len

    def extract_assignment(self, key):
        data = self.expands[key]
        e_len = len(data)
        self.expands.pop(key)
        self.volume -= e_len
        return data

    def delta_sqr(self, key, values):
        if isinstance(values, int):
            cntr = values
        else:
            cntr = len(values)
            if self.id in values:
                cntr -= 1
        return (self.volume + cntr) ** 2 - self.cost()

    def __str__(self):
        return self.expands.__str__()

    def __unicode__(self):
        return self.expands.__str__()

    def __repr__(self):
        return self.expands.__str__()


class dist_tracker:
    def __init__(self):
        self.other_idx = -1
        self.other_send_list = []
        self.send_list = []

    def set(self, other_idx: int, other_send_list: list, send_list: list):
        self.other_idx = other_idx
        self.other_send_list = other_send_list
        self.send_list = send_list

    def reset(self):
        self.other_send_list.clear()
        self.send_list.clear()
        self.other_idx = -1


def write_results(arr):
    sys.stdout = open(f"out/reddit_{arr[4]}cores.txt", "w")
    print("Vertex Count:", arr[0])
    print("Edge count: %d" % arr[1])
    print(f"Algorithm execution time: {arr[2]} seconds")
    print(f"Algorithm and data parsing execution time: {arr[3]} seconds")
    print("Core Count %d" % arr[4])
    print("Number of extend operations: %d" % arr[5])
    print("Number of non-extended operations: %d" % arr[6])
    print("Min Sum Square: %d, Initial: %d" % (arr[7], arr[8]))
    print("Min Sum Square without delta: %d" % arr[9])
    print("Highest Volume %d, Initial: %d" % (arr[10], arr[11]))
    sys.stdout.close()


def save_basic_plot(x, y1, xlabel, ylabel, title):
    fig, ax = plt.subplots()
    ax.plot(x, y1)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    plt.savefig(f"out/{title}.png")


def save_2basic_plot(x, y1, y2, xlabel, ylabel, title, y1name, y2name, yscale=None):
    fig, ax = plt.subplots()
    ax.plot(x, y1, label=y1name)
    ax.plot(x, y2, label=y2name)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    if yscale is not None:
        ax.set_yscale(yscale)
    ax.legend()
    plt.savefig(f"out/{title}.png")


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):
        col_name = openpyxl.utils.cell.get_column_letter(min_col + i)
        value = column_widths[i] + 1
        ws.column_dimensions[col_name].width = value


def create_excel(core_cnt, data, datasets):
    wb = Workbook()
    ws = wb.active

    DEFAULT_FONT.sz = 10
    # _font = Font(sz=10)
    # {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"] = "Dataset"
    interval = len(data) / len(datasets)

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws["B1"] = "Vertex\nCnt"
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws["C1"] = "Edge\nCnt"
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    ws["D1"] = "Time (s)"
    ws["D2"] = "Algo"
    ws["E2"] = "Total"
    ws.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
    ws["F1"] = "Cores"
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    ws["G1"] = "Reassignment"
    ws["G2"] = "Cnt"
    ws["H2"] = "Non-R. Cnt"
    ws["I2"] = "Vol"
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)
    ws["J1"] = "Vol (Init/Opt) Ratio"
    ws["J2"] = "Square Sum"
    ws["K2"] = "Max"
    ws["L2"] = "CV"
    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=17)
    ws["M1"] = "Vol"
    ws["M2"] = "Avg"
    ws["N2"] = "I CV"
    ws["O2"] = "O CV"
    ws["P2"] = "I Min"
    ws["Q2"] = "O Min"
    ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=24)
    ws["R1"] = "Degree"
    ws["R2"] = "Cnt"
    ws["S2"] = "Avg"
    ws["T2"] = "Total"
    ws["U2"] = "Max"
    ws["V2"] = "Min"
    ws["W2"] = "CV"
    ws["X2"] = "High %10"
    ws.merge_cells(start_row=1, start_column=25, end_row=2, end_column=25)
    ws["Y1"] = "Avg\nFwd Vol"

    # ws["X2"] = "Low %20 Perc."

    for i in range(len(data)):
        d = data[i]
        for j in range(len(d)):
            ws.cell(row=i + 3, column=j + 2, value=d[j])
    adjust_column_width_from_col(ws, 2, 1, ws.max_column)
    adjust_column_width_from_col(ws, 1, 1, 1)

    double = borders.Side(border_style="thin")

    for i in range(ws.max_column):
        ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(1)].alignment = openpyxl.styles.Alignment(
            horizontal="center")
        for j in range(1, ws.max_row):
            cell = ws[openpyxl.utils.cell.get_column_letter(i + 1) + str(j + 1)]
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="right", vertical="center")
            if j % interval == 1:
                cell.border = Border(bottom=double)

    # for j in range(1, ws.max_row + 1):
    #     end_cell = ws.cell(row=j, column=ws.max_column)
    #     end_cell.border = Border(right=double, bottom=double)
    # row = ws.row_dimensions[2]
    for i in range(len(datasets)):
        d = datasets[i]
        ws.merge_cells(start_row=interval * i + 3, start_column=1, end_row=interval * (i + 1) + 2, end_column=1)
        cell = ws.cell(row=interval * i + 3, column=1, value=d)
        cell.border = Border(right=double)

    wb.save(f'out/{core_cnt}cores.xlsx')


class Assigment(Enum):
    N_MINUS_1_TO_1 = 1
    HALF_SPLIT = 2
    VOL_EQ_SPLIT = 3
    NO_REASSIGNMENT = 0


def assignment_cost_test(a, b, set_size):
    diff = a - b
    if math.fabs(diff) < set_size:
        if diff > 0:
            other_vol = int(math.ceil((set_size - diff) / 2))
            owner_vol = int(math.floor((set_size + diff) / 2))
        else:
            other_vol = int(math.floor((set_size - diff) / 2))
            owner_vol = int(math.ceil((set_size + diff) / 2))
    else:
        other_vol = set_size - 1
        owner_vol = 1
    print(f"other: {other_vol}, owner: {owner_vol}")


def get_vol_eq_split_vols(a, b, set_size):
    diff = a - b
    other_vol = int(math.ceil((set_size - diff) / 2))
    owner_vol = int(math.floor((set_size + diff) / 2))
    if owner_vol == 0:
        owner_vol += 1
        other_vol -= 1
    elif other_vol == 0:
        owner_vol -= 1
        other_vol += 1
    return other_vol, owner_vol


def assignment_cost(other_data: DestData, core_data: DestData, key: int, send_set: set):
    diff = other_data.volume - core_data.volume
    set_size = len(send_set)
    if np.abs(diff) < set_size:
        other_vol, owner_vol = get_vol_eq_split_vols(other_data.volume, core_data.volume, set_size)
        return other_data.delta_sqr(key, other_vol) + core_data.delta_sqr(key, owner_vol), Assigment.VOL_EQ_SPLIT
    else:
        return other_data.delta_sqr(key, send_set) + core_data.delta_sqr(key, 1), Assigment.N_MINUS_1_TO_1


class MetricTracker:
    def __init__(self):
        self.cost = 0
        self.reassign_cnt = 0
        self.non_reassign_cnt = 0
        self.reassign_vol = 0
        self.non_reassignment_vol = 0

    def set_vol_eq_split(self, o_vol, c_vol):
        self.reassign_cnt += 1
        self.non_reassignment_vol += c_vol
        self.reassign_vol += o_vol

    def set(self, atype: Assigment, set_size: int):
        match atype:
            case Assigment.NO_REASSIGNMENT:
                self.non_reassign_cnt += 1
                self.non_reassignment_vol += set_size
            case Assigment.N_MINUS_1_TO_1:
                self.reassign_cnt += 1
                self.non_reassignment_vol += 1
                self.reassign_vol += set_size - 1
            case _:
                raise Exception("This assignment type is not supported")
